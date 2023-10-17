import firebase_admin
from firebase_admin import credentials, firestore
from openpyxl import Workbook
from openpyxl import load_workbook
import csv

# Inicializar o Firebase Admin SDK
cred = credentials.Certificate("c:\\Users\\ENZOVITORIANOPUTTINI\\Desktop\\Projeto_TCC\\sdk.json")
firebase_admin.initialize_app(cred)

# Inicializar o Firestore
db = firestore.client()
print("O Banco de dados foi acessado")

def salvar_firestore_em_xlsx():
    # Recupere os dados do Firestore
    alunos = db.collection("matscan").get()

    try:
        workbook = load_workbook("alunos.xlsx")
    except FileNotFoundError:
        workbook = Workbook()

    sheet = workbook.active

    if "Nome Completo" not in [cell.value for cell in sheet[1]]:
        sheet["A1"] = "Classe"
        sheet["B1"] = "Nome Completo"
        sheet["C1"] = "RA"
        sheet["D1"] = "QR Code"
        sheet["E1"] = "Email"
        sheet["F1"] = "Nome do Responsavel"

    # Crie um conjunto (set) para armazenar os nomes de alunos já existentes na planilha
    nomes_na_planilha = set(sheet.cell(row=i, column=2).value for i in range(2, sheet.max_row + 1))

    for aluno in alunos:
        classe = aluno.get("className").upper()
        name = aluno.get("fullName").upper()
        qr_code = aluno.get("qrCode")
        ra = aluno.get("ra")
        email = aluno.get("email").lower()

        # Verifique se o name do aluno já existe na planilha antes de adicioná-lo
        if name not in nomes_na_planilha:
            # Adicione os dados do aluno à planilha
            nova_linha = [classe, name, ra, qr_code, email]# quando tiver os campos email e respName no banco, não esquecer de adicionar as variaveis nesta linha
            sheet.append(nova_linha)
            nomes_na_planilha.add(name)

    print("Arquivo xlsx criado")

    workbook.save("alunos.xlsx")

def xlsx_para_csv():
    # Abra o arquivo XLSX
    try:
        workbook = load_workbook("alunos.xlsx")
    except FileNotFoundError:
        workbook = Workbook()

    # Selecione a planilha que você deseja converter em CSV
    planilha = workbook.active  # Ou use workbook["nomedaplanilha"] para selecionar uma planilha específica

    # Crie um arquivo CSV para escrever os dados
    arquivo_csv = "alunos.csv"
    with open(arquivo_csv, "w", newline="") as arquivo:
        escritor_csv = csv.writer(arquivo)

        # Itere pelas linhas da planilha e escreva-as no arquivo CSV
        for linha in planilha.iter_rows(values_only=True):
            escritor_csv.writerow(linha)

    print(f"Conversão concluída. Os dados foram salvos em {arquivo_csv}")

salvar_firestore_em_xlsx()
xlsx_para_csv()
